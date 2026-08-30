"""
Day-1 deliverable: the merged label table, all layers joined per compound.

Sources, all on disk:
  - Refinement-Table.xlsx: Correction verdicts, auto and manual Rwp and weight
    fractions, ordering notes, rival phases with ICSD codes (the Correction layer)
  - data/labels/correction_verdicts.csv (the four plus the removed one; cross-check)
  - data/labels/prx_table1_errors.csv (critics' per-compound error flags)
  - data/labels/prx_table3.csv (critics' rival assignments; 42 of 43 rows)
  - results/structure_scan.csv (our internal-consistency scan)
  - data/cifs/{Automated,Manual}_Refinement_Results (file paths, resolved and hashed)

Output: data/labels/merged_labels.csv, one row per compound (42 deposited plus the
removed Zn2Cr3FeO8), and the H5 statistics (Rwp against verdict) on stdout.

H5 preregistration, stated before this was ever computed: fit quality
is expected NOT to track the verdict; a near-zero relationship is the anticipated
and publishable outcome.
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "cifs" / "Refinement-Table.xlsx"
OUT = ROOT / "data" / "labels" / "merged_labels.csv"

def repo_relative(uri: str) -> str:
    """Paths in committed tables are relative to the repo, so the table is the same
    file on every machine and carries nobody's home directory."""
    try:
        return str(Path(uri).resolve().relative_to(ROOT))
    except ValueError:
        return str(uri)


RWP_RE = re.compile(r"Rwp\s*=\s*([\d.]+)")
FRAC_RE = re.compile(r"\[([\d.]+)%")
ICSD_RE = re.compile(r"ICSD\s*#?\s*(\d+)")


def resolve_cif(stage: str, name: str) -> Path:
    """Find a compound's deposited file, by name first and by COMPOSITION when
    the name misses.

    The deposit and the Correction's spreadsheet do not always spell a formula
    with its elements in the same order: the spreadsheet's Ba6Ta2Na2V2O17 is
    the deposit's folder Ba6Na2Ta2V2O17, and the folder Y3Ga3In2O12 contains a
    file named Y3In2Ga3O12.cif. Building the path from the spreadsheet name
    alone therefore missed two of the 42 deposited pairs, and one of them,
    Ba6Ta2Na2V2O17, is in the evaluation population - so every check on it was
    skipped for want of evidence and its ledger row read as a verdict rather
    than as an absence.

    Section 5.3 of the benchmark already described this fallback as the way the
    pair is resolved; it did not exist in the code until 26 August 2026. The
    composition comparison is the same normalisation used to join the label
    tables to each other.
    """
    root = ROOT / "data" / "cifs" / stage
    direct = root / name / f"{name}.cif"
    if direct.exists():
        return direct
    want = norm_formula(name)
    if not root.exists():
        return direct
    for cand in sorted(root.rglob("*.cif")):
        if norm_formula(cand.stem) == want:
            return cand
    return direct


def norm_formula(name: str) -> str:
    """Composition-normalised join key, tolerant of permuted formulas."""
    try:
        from pymatgen.core import Composition
        return Composition(name).reduced_formula
    except Exception:
        return name.strip()


def parse_xlsx() -> tuple[list[dict], int]:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    records, section, current, n_offline = [], "main", None, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        a = (str(row[0]).strip() if row[0] is not None else "")
        b = (str(row[1]).strip() if row[1] is not None else "")
        if a and not b:
            low = a.lower()
            if "inconclusive" in low:
                section = "inconclusive"
            elif "offline" in low:
                section = "offline"
            else:
                section = a
            continue
        if a and b:
            if re.fullmatch(r"[*\s]+", a):     # footnote rows, not compounds
                current = None
                continue
            if section == "offline":
                n_offline += 1
            current = {
                "compound": a.replace("\n", " ").strip(),
                "section": section,
                "conclusions": b,
                "auto_text": str(row[2] or ""),
                "man_text": str(row[3] or ""),
                "notes": str(row[4] or ""),
                "rivals": [],
            }
            records.append(current)
        # lattice or rival continuation rows attach to the current compound
        if current is not None and not a:
            phase = str(row[5] or "").strip()
            if phase:
                current["rivals"].append(
                    {"phase": phase.replace("\n", " "),
                     "sg": str(row[6] or "").strip(),
                     "icsd": (ICSD_RE.search(phase).group(1)
                              if ICSD_RE.search(phase) else "")})
    return records, n_offline


def first_float(rx: re.Pattern, text: str) -> float | None:
    m = rx.search(text)
    return float(m.group(1)) if m else None


def load_csv(path: Path, key: str) -> dict[str, dict]:
    out = {}
    with path.open() as fh:
        for row in csv.DictReader(r for r in fh if not r.startswith("#")):
            out[norm_formula(row[key])] = row
    return out


def main() -> int:
    records, n_offline = parse_xlsx()
    print(f"xlsx: {len(records)} compounds in main plus inconclusive sections, "
          f"{n_offline} offline rows (excluded)")

    prx_err = load_csv(ROOT / "data/labels/prx_table1_errors.csv", "compound")
    prx_t3 = load_csv(ROOT / "data/labels/prx_table3.csv", "sample")
    corr = load_csv(ROOT / "data/labels/correction_verdicts.csv", "compound")
    # REFUSE RATHER THAN BLANK THE COLUMNS. The scan is this project's own
    # internal-consistency pass over the deposited files; it lands in results/,
    # which is gitignored, so it does not travel with the benchmark. Every read
    # of it below is a .get with a default, so a missing scan would quietly
    # produce a merged table whose scan_auto_sg, scan_auto_disordered and
    # scan_mixed_sites columns are empty - and the shipped key has them
    # populated on 42 of 43 rows. That is a silently different answer key,
    # which is the failure this file exists to prevent.
    scan_path = ROOT / "results/structure_scan.csv"
    if not scan_path.exists():
        raise SystemExit(
            f"REFUSING TO BUILD: {scan_path} is absent.\n"
            f"It is this project's structure scan, written to the gitignored "
            f"results/ tree and not shipped\n"
            f"with the benchmark. Building without it would write a "
            f"merged_labels.csv whose scan_auto_sg,\n"
            f"scan_auto_disordered and scan_mixed_sites columns are empty, "
            f"where the shipped key has them\n"
            f"on 42 of 43 rows - a different answer key, produced silently. "
            f"See known-defects.md 8i.")
    scan = load_csv(scan_path, "compound")

    from claim import resolve_ref

    rows = []
    for rec in records:
        name = rec["compound"]
        key = norm_formula(name)
        concl = rec["conclusions"].replace("\n", " ")
        sm = re.search(r"[Ss]tructure:\s*([A-Za-z]+)", concl)
        cm = re.search(r"[Cc]omposition:\s*([A-Za-z]+)", concl)
        auto_cif = resolve_cif("Automated_Refinement_Results", name)
        man_cif = resolve_cif("Manual_Refinement_Results", name)
        a_ref = resolve_ref(auto_cif, "cif")
        m_ref = resolve_ref(man_cif, "cif")
        t3 = prx_t3.get(key, {})
        sc = scan.get(key, {})
        rows.append({
            "compound": name,
            "section": rec["section"],
            "correction_structure": (sm.group(1).lower() if sm else ""),
            "correction_composition": (cm.group(1).lower() if cm else ""),
            "correction_verdict": corr.get(key, {}).get("correction_verdict",
                {"main": "confirmed", "inconclusive": "inconclusive",
                 "offline": "offline, not adjudicated"}.get(rec["section"],
                                                            rec["section"])),
            "rwp_auto": first_float(RWP_RE, rec["auto_text"]),
            "rwp_manual": first_float(RWP_RE, rec["man_text"]),
            "frac_auto_pct": first_float(FRAC_RE, rec["auto_text"]),
            "frac_manual_pct": first_float(FRAC_RE, rec["man_text"]),
            "prx_e1": prx_err.get(key, {}).get("e1_poor_fit", ""),
            "prx_e2": prx_err.get(key, {}).get("e2_cif_differs", ""),
            "prx_e3": prx_err.get(key, {}).get("e3_no_ordering_evidence", ""),
            "prx_e4": prx_err.get(key, {}).get("e4_already_known", ""),
            "prx_proposed_sym": t3.get("proposed_sym", ""),
            "prx_indexed_sym": t3.get("indexed_sym", ""),
            "prx_candidates": t3.get("icsd_candidates", ""),
            "prx_icsd_codes": t3.get("icsd_codes", ""),
            "prx_uncertain": t3.get("uncertain", ""),
            "scan_auto_sg": sc.get("auto_sg", ""),
            "scan_auto_disordered": sc.get("auto_cation_disordered", ""),
            "scan_mixed_sites": sc.get("auto_mixed_sites", ""),
            "correction_rivals": "; ".join(
                f"{r['phase']}" + (f" [ICSD {r['icsd']}]" if r["icsd"] else "")
                for r in rec["rivals"][1:]),   # first lattice row is the claim itself
            "auto_cif": repo_relative(a_ref.uri), "auto_cif_sha256": a_ref.sha256 or "",
            "auto_cif_resolved": a_ref.resolved,
            "man_cif": repo_relative(m_ref.uri), "man_cif_sha256": m_ref.sha256 or "",
            "man_cif_resolved": m_ref.resolved,
            "ordering_note": ("Ordering:" in rec["notes"]),
        })

    # the removed compound, carried as a row with unresolved evidence (absence is data)
    zn = corr.get(norm_formula("Zn2Cr3FeO8"), {})
    zne = prx_err.get(norm_formula("Zn2Cr3FeO8"), {})
    rows.append({**{k: "" for k in rows[0]},
                 "compound": "Zn2Cr3FeO8", "section": "removed",
                 "correction_verdict": zn.get("correction_verdict", "removed"),
                 "prx_e1": zne.get("e1_poor_fit", ""),
                 "prx_e2": zne.get("e2_cif_differs", ""),
                 "prx_e3": zne.get("e3_no_ordering_evidence", ""),
                 "prx_e4": zne.get("e4_already_known", ""),
                 "auto_cif_resolved": False, "man_cif_resolved": False})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"wrote {OUT} ({len(rows)} rows)")

    # unmatched joins
    for label, layer in (("prx_table3", prx_t3), ("scan", scan)):
        missing = [r["compound"] for r in rows[:-1]
                   if norm_formula(r["compound"]) not in layer]
        if missing:
            print(f"  NOT MATCHED in {label}: {missing}")

    # ---------------- H5: does fit quality track the verdict?
    from scipy.stats import mannwhitneyu, spearmanr
    for col in ("rwp_auto", "rwp_manual"):
        pairs = [(r[col], r["correction_verdict"]) for r in rows[:-1]
                 if isinstance(r[col], float)]
        inc = [v for v, verd in pairs if verd == "inconclusive"]
        con = [v for v, verd in pairs if verd != "inconclusive"]
        if len(inc) >= 3 and con:
            u, p = mannwhitneyu(inc, con, alternative="two-sided")
            rho, ps = spearmanr([v for v, _ in pairs],
                                [1 if verd == "inconclusive" else 0 for _, verd in pairs])
            print(f"\nH5 [{col}] n={len(pairs)} (inconclusive {len(inc)}, other {len(con)})")
            print(f"  inconclusive Rwp: {sorted(inc)}")
            print(f"  Mann-Whitney U={u:.1f} p={p:.3f}; Spearman rho={rho:+.3f} p={ps:.3f}")
    print("\nH5 preregistered expectation: near zero. Report either way.")

    # ---------------- the panel-disagreement accounting, from the corrected lists
    e3 = {r["compound"] for r in rows if r["prx_e3"] == "1"}
    inc = {r["compound"] for r in rows if r["correction_verdict"] == "inconclusive"}
    conf = {r["compound"] for r in rows if r["correction_verdict"] == "confirmed"}
    off = {r["compound"] for r in rows if "offline" in r["correction_verdict"]}
    print(f"\nPanel disagreement (printed Table I against the Correction):")
    print(f"  E3 total {len(e3)}; inconclusive-and-E3 {sorted(e3 & inc)}")
    print(f"  CONFIRMED-and-E3 (the disagreement-test positives): {len(e3 & conf)}")
    print(f"  E3 members outside adjudication: removed "
          f"{sorted(e3 - conf - inc - off)}, offline {sorted(e3 & off)}")

    e2 = {r["compound"] for r in rows if r["prx_e2"] == "1"}
    scan10 = {r["compound"] for r in rows if r["scan_auto_disordered"] == "1"}
    print(f"\nE2 against our scan's disorder flags:")
    print(f"  both ({len(e2 & scan10)}): {sorted(e2 & scan10)}")
    print(f"  E2 only: {sorted(e2 - scan10)}")
    print(f"  scan only: {sorted(scan10 - e2)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
